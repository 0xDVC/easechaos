import pandas as pd
import pytest
from api.extract.extract_exam_table import (
    _find_header_row,
    _resolve_column_name,
    _normalize_columns,
    _convert_exam_dates,
    get_exam_timetable,
    PERIOD_MAPPING,
)


class TestFindHeaderRow:
    def test_finds_header_with_date_and_class(self):
        df = pd.DataFrame({
            0: ["junk", "DATE", "data"],
            1: ["junk", "CLASS", "data"],
        })
        assert _find_header_row(df) == 1

    def test_finds_header_with_session_alias(self):
        df = pd.DataFrame({
            0: ["junk", "DATE", "data"],
            1: ["junk", "SESSION", "data"],
        })
        assert _find_header_row(df) == 1

    def test_raises_when_header_missing(self):
        df = pd.DataFrame({"A": ["foo", "bar"]})
        with pytest.raises(ValueError, match="Could not find exam timetable header row"):
            _find_header_row(df)


class TestResolveColumnName:
    def test_exact_match(self):
        cols = pd.Index(["COURSE NO", "CLASS", "DATE"])
        assert _resolve_column_name(cols, ["COURSE NO"]) == "COURSE NO"

    def test_alias_match(self):
        cols = pd.Index(["CRS CODE", "CLASS", "DATE"])
        assert _resolve_column_name(cols, ["COURSE NO", "CRS CODE"]) == "CRS CODE"

    def test_no_match(self):
        cols = pd.Index(["A", "B", "C"])
        assert _resolve_column_name(cols, ["COURSE NO"]) is None


class TestNormalizeColumns:
    def test_renames_all_known_columns(self):
        df = pd.DataFrame(columns=["CRS CODE", "COURSE TITLE", "LECTURE HALL"])
        result = _normalize_columns(df)
        assert "COURSE NO" in result.columns
        assert "COURSE NAME" in result.columns
        assert "LECTURE HALL" in result.columns

    def test_leaves_unknown_columns_unchanged(self):
        df = pd.DataFrame(columns=["SOME_RANDOM_COL"])
        result = _normalize_columns(df)
        assert "SOME_RANDOM_COL" in result.columns


class TestConvertExamDates:
    def test_serial_numbers(self):
        s = pd.Series([45000, 45001])
        result = _convert_exam_dates(s)
        assert result.dtype.kind == "M"

    def test_string_dates(self):
        s = pd.Series(["2024-01-15", "2024-01-16"])
        result = _convert_exam_dates(s)
        assert result.dtype.kind == "M"

    def test_coerce_invalid_to_na(self):
        s = pd.Series(["not-a-date", "also-invalid"])
        result = _convert_exam_dates(s)
        assert result.isna().all()


class TestGetExamTimetable:
    def test_returns_filtered_dataframe(self, tmp_path):
        import openpyxl
        wb = openpyxl.Workbook()
        ws = wb.active
        assert(ws is not None)
        ws.append(["junk", "col"])
        ws.append(["DATE", "CLASS", "PERIOD", "COURSE NAME"])
        ws.append(["45000", "CE 4A", "M", "Math 101"])
        ws.append(["45001", "CE 4B", "A", "Phy 101"])
        ws.append(["45002", "ME 5A", "M", "Mech 301"])
        path = tmp_path / "exam.xlsx"
        wb.save(path)

        result = get_exam_timetable(str(path), "CE 4")
        assert len(result) == 2
        assert all(result["CLASS"].str.startswith("CE 4"))
        assert "START" in result.columns
        assert "END" in result.columns
        assert list(result["START"]) == ["7:00 AM", "11:00 AM"]

    def test_empty_when_no_matching_class(self, tmp_path):
        import openpyxl
        wb = openpyxl.Workbook()
        ws = wb.active
        assert(ws is not None)
        ws.append(["DATE", "CLASS", "PERIOD", "COURSE NAME"])
        ws.append(["45000", "ME 5A", "M", "Mech 301"])
        path = tmp_path / "exam.xlsx"
        wb.save(path)

        result = get_exam_timetable(str(path), "CE 4")
        assert len(result) == 0

    def test_drops_no_column(self, tmp_path):
        import openpyxl
        wb = openpyxl.Workbook()
        ws = wb.active
        assert(ws is not None)
        ws.append(["DATE", "CLASS", "PERIOD", "COURSE NAME", "NO"])
        ws.append(["45000", "CE 4A", "M", "Math 101", "1"])
        path = tmp_path / "exam.xlsx"
        wb.save(path)

        result = get_exam_timetable(str(path), "CE 4")
        assert "NO" not in result.columns
